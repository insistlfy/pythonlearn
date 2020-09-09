# !usr/bin/env python
# -*- coding:utf-8 _*-
"""
@author:lfy
@file: excel_util.py
@time: 2020-09-07 下午3:38
@description: excel_util
"""

import openpyxl
import re
import traceback
import os


class Excel:

    def change_data(self, file_dir, source, target):
        wb = openpyxl.load_workbook(file_dir)
        num = 0

        # 获取所有的sheet页
        sheet_list = wb.sheetnames
        print(wb.sheetnames)

        for sheet in sheet_list:
            ws = wb.worksheets[num]
            num = num + 1

    if __name__ == '__main__':
        change_data('', '', '')
